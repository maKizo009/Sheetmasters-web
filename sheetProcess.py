from fileinput import filename
from flask import app, redirect, url_for, request
from werkzeug.utils import secure_filenameimport pandas as pd
import csv
import os
from openpyxl import load_workbook
from shutil import copyfile
from upload import upload_file

@app.route('/sheetProcess')
def process_file():
    global processing_completed

    filename = upload_file()
    with open(filename, "r") as f_entrada:
        leitor = csv.reader(f_entrada, delimiter=";")

        #Pegando o nome da estação
        primeira_linha = f_entrada.readline().rstrip()
        print(primeira_linha)

        #Tranformando os dados da planilha em um Dataframe
        leitor = pd.DataFrame(leitor)

    #Ecluindo as partes inúteis da planilha, localizada nas primeiras 11 linhas
        leitor = leitor.iloc[11:]

    #Renomeando e reordenando as colunas da forma que eu quero
        leitor = leitor.rename(columns={0: "Data", 1: "Max", 2: "Min", 3: "Precipitação"})
        leitor = leitor.reindex(columns=["Data", "Min", "Max"])
    #Tranformando os dados de string para float
        leitor["Max"] = pd.to_numeric(leitor["Max"], errors="coerce")#.astype(float)
        leitor["Min"] = pd.to_numeric(leitor["Min"], errors="coerce")#.astype(float)

    #Criando um pequeno resumo da estação
        linha = leitor["Max"].idxmin()
        colunas = ["Data", "Min", "Max"]
        resultado = leitor.loc[linha, colunas]

        linha2 = leitor["Min"].idxmin()
        colunas2 = ["Data", "Min", "Max"]
        resultado2 = leitor.loc[linha2, colunas2]

        linha3 = leitor["Max"].idxmax()
        colunas3 = ["Data", "Min", "Max"]
        resultado3 = leitor.loc[linha3, colunas3]

        linha4 = leitor["Min"].idxmax()
        colunas4 = ["Data", "Min", "Max"]
        resultado4 = leitor.loc[linha4, colunas4]

        leitor["Data"] = pd.to_datetime(leitor["Data"])

        leitor["Dia"] = leitor["Data"].dt.day
        leitor["Mes"] = leitor["Data"].dt.month
        leitor["Ano"] = leitor["Data"].dt.year

    caminhoOriginal = "Modelo para normais climatológicas 1991-2020.xlsx"
    novoCaminho = (primeira_linha[6:]+"_BDMEP.xlsx")

    copyfile(caminhoOriginal, novoCaminho)

        #Imprimindo os resultados na tela
    # dados1 = ("A estação de {}, registrou sua menor máxima em {}, com {} graus, e a maior em {} com {}.".format(primeira_linha, resultado["Data"], resultado["Max"], resultado3["Data"], resultado3["Max"]))
    # dados2 = ("Nas mínimas, a menor foi {} graus em {}, e a maior foi {}, em {}.".format(resultado2["Min"], resultado2["Data"], resultado4["Min"], resultado4["Data"]))

    # print(dados1, dados2)
    # print("Sua planilha está sendo automatizada...")

    book = load_workbook("Modelo para normais climatológicas 1991-2020.xlsx")

    for ano in leitor["Ano"].dropna().unique():
        dadosAno = leitor[leitor["Ano"]== ano]
        organizado = pd.DataFrame()

        for mes in range(1, 13):
            dadosMes = dadosAno[dadosAno["Mes"] == mes]
            minimas = dadosMes[["Dia", "Min"]].set_index("Dia")
            maximas = dadosMes[["Dia", "Max"]].set_index("Dia")
            minimas.columns = [f"Minima{mes}"]
            maximas.columns = [f"Máxima{mes}"]
            organizado.index = organizado.index.astype(str)
            inicio = pd.to_datetime(f'{int(ano)}-{int(mes):02d}-01')
            fim = inicio + pd.offsets.MonthEnd(0)
            dias = pd.date_range(inicio, fim)
            dias = dias.day
            minimas = minimas.reindex(dias)
            maximas = maximas.reindex(dias)
            organizado = organizado.reset_index(drop=True)
            minimas = minimas.reset_index(drop=True)
            maximas = maximas.reset_index(drop=True)
            organizado = pd.concat([organizado, minimas, maximas], axis=1)

        sheet_name = str(int(ano))
        if sheet_name in book.sheetnames:
            ws = book[sheet_name]
        else:
            ws = book.create_sheet(sheet_name)

        for col, col_data in enumerate(organizado.values.T, start=2):
            for row, value in enumerate(col_data, start=3):
                if pd.isna(value):
                    value = "-"
                ws.cell(row=row, column=col, value=value)


    book.save(novoCaminho)
    processing_completed = True

    return redirect(url_for('download_file', filename=novoCaminho))
   